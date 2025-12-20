import os
from typing import Optional, cast
import pandas as pd
from openpyxl import load_workbook


def get_xls_text(content):
    excel_file = pd.ExcelFile(content, engine="xlrd")
    texts = []
    for excel_sheet_name in excel_file.sheet_names:
        df = excel_file.parse(sheet_name=excel_sheet_name)
        df.dropna(how="all", inplace=True)

        for _, row in df.iterrows():
            page_content = []
            for k, v in row.items():
                if pd.notna(v):
                    page_content.append(f'"{k}":"{v}"')
            texts.append(";".join(page_content))
    return "\n".join(texts)


def get_xlsx_text(content):
    texts = []
    wb = load_workbook(content, data_only=True)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        data = sheet.values
        try:
            cols = next(data)
        except StopIteration:
            continue
        df = pd.DataFrame(data, columns=cols)

        df.dropna(how="all", inplace=True)

        for index, row in df.iterrows():
            page_content = []
            for col_index, (k, v) in enumerate(row.items()):
                if pd.notna(v):
                    cell = sheet.cell(
                        row=cast(int, index) + 2, column=col_index + 1
                    )  # +2 to account for header and 1-based index
                    if cell.hyperlink:
                        value = f"[{v}]({cell.hyperlink.target})"
                        page_content.append(f'"{k}":"{value}"')
                    else:
                        page_content.append(f'"{k}":"{v}"')
            texts.append(";".join(page_content))
    return "\n".join(texts)


def get_csv_text(content):
    texts = []
    df = pd.read_csv(content, on_bad_lines="skip")

    for i, row in df.iterrows():
        content = ";".join(f"{col.strip()}: {str(row[col]).strip()}" for col in df.columns)
        texts.append(content)
    return "\n".join(texts)
